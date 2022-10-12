from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import traceback
import configparser
import glob
import openpyxl
import os
import ntpath
import pandas as pd
import json
from datetime import datetime, date

# config = configparser.ConfigParser()
# config.read('env.ini')

app = Flask(__name__, static_url_path="", static_folder="public") # 指定public目錄為static folder
app.config.from_object(__name__)
CORS(app, resources={r'/*': {'origins': '*'}})

# 檢查目錄是否存在
if not os.path.isdir("data"):
    os.mkdir("data")

# 接收nb watch dog上傳的資料
@app.route('/upload', methods = ['POST'])
def upload_file():
    try:
        if request.method == 'POST':
            file = request.files['filedata']   # gives you a FileStorage
            filename = secure_filename(file.filename)
            file.save(os.path.join("data/", filename))
            response = {"status": 'success'}
            return jsonify(response)
    except Exception:
        print(traceback.format_exc())

# 接收設備維修頁面上傳的資料
@app.route('/repairRecord', methods = ['POST'])
def repair_record():
    try:
        if request.method == 'POST':
            #datename = 'rent'+datetime.today().strftime('%Y-%m-%d')+'.csv'
            body = request.get_data(as_text=True)  # 取得收到的訊息內容
            json_data = json.loads(body) # json 格式化訊息內容
            print(json_data)
            df = pd.json_normalize(json_data)
            # 更新最新維修狀態
            if os.path.exists(os.path.join("manage/repair.csv")):
                repairDf = pd.read_csv ("manage/repair.csv")
                # 只保留最新一筆
                repairDf = repairDf[~repairDf.nbNumber.str.contains(json_data['nbNumber'])] # 移除符合條件的row data
                result = pd.concat([repairDf, df])
                result.to_csv("manage/repair.csv",encoding='utf_8_sig',index=False)
            else:
                df.to_csv("manage/repair.csv",encoding='utf_8_sig',index=False)
            response = {"status": 'success'}
            return jsonify(response)
    except Exception:
        print(traceback.format_exc())

# 列出所有維修清單
@app.route('/repairInfo', methods = ['GET'])
def repair_info():
    try:
        if request.method == 'GET':
            # 整理所有維修表單
            repairData = {} # 以筆電編號當 key
            if os.path.exists(os.path.join("manage/repair.csv")):
                with open(os.path.join("manage/repair.csv"),"r",encoding="utf-8") as f:
                    lines = f.readlines()
                    for line in lines:
                        if line.split(",")[1] == "nbNumber":
                            continue
                        repairData[line.split(",")[1]] = line.split(",")
            
             # 整理所有上傳資料
            nbList = glob.glob("data/*_info.txt")
            
            # ------------------------------以上傳資料為主去走訪-------------------------------
            nameList = []
            for nb in nbList:
                nameList.append(ntpath.basename(nb.split("_info")[0])) # 取得所有有上傳資料的電腦名稱(=筆電編號 nbNumber)
            nbInfo = {}
            for name in nameList: # 走訪每台電腦取得所需資訊
                info = {}
                info["nbName"] = name
                info["nbNumber"] = name
                with open(os.path.join("data/", name + '_info.txt')) as f:
                    lines = f.readlines()
                    values = lines[1].split(",")
                    info["os"] =  values[13].replace('"','') + ' ' + values[14].replace('"','') # 取得硬體資訊
                if os.path.exists(os.path.join("data/", name + '_logon.txt')):
                    with open(os.path.join("data/", name + '_logon.txt')) as f:
                        lines = f.readlines()
                        user = ""
                        time = 0
                        for line in lines:
                            if len(line.split()) <= 6:
                                continue
                            else:
                                user = line.split()[0].strip()
                                user = user[1:len(user)]
                                time = line.split()[5].strip()+' '+line.split()[6].strip()+' '+line.split()[7].strip()
                        info["user"] = user # 取得最後登入者
                        info["logonTime"] = time # 取得最後登入時間
                if os.path.exists(os.path.join("data/", name + '_product.txt')):
                    productList = []
                    with open(os.path.join("data/", name + '_product.txt'), encoding='utf-16') as f:
                        lines = f.readlines()
                        for line in lines:
                            if ('Name' in line) and ('Version' in line):
                                continue
                            if len(line.split(",")) == 3:
                                if line.split(",")[1] and line.split(",")[2]:
                                    item = {}
                                    item["name"] = line.split(",")[1]
                                    item["version"] = line.split(",")[2].replace('\n',"").replace('\t',"")
                                    productList.append(item["name"]+":"+item["version"])       
                        info["product"] = productList # 取得軟體清單:版本資訊
                nbInfo[name] = info

            # ------------------------------維修表單整合上傳資料-------------------------------
            response = []
            for name in repairData.keys():
                repairInfo = {}
                repairInfo['fab'] = repairData[name][0]
                repairInfo['nbNumber'] = repairData[name][1]
                repairInfo['formNo'] = repairData[name][2]
                repairInfo['repairStatus'] = repairData[name][3]
                repairInfo['overDate'] = repairData[name][4]
                repairInfo['registerDate'] = repairData[name][5]
                repairInfo['remark'] = repairData[name][6]
                if name in nameList:
                    repairInfo['os'] = nbInfo[name]['os']
                    repairInfo['logonTime'] = nbInfo[name]['logonTime']
                else:
                    repairInfo['os'] = '-'
                    repairInfo['logonTime'] = '-'
                response.append(repairInfo)
            return jsonify(response)

            #print(repairData)
            #if os.path.exists(os.path.join("manage/repair.csv")):
            #    repairDf = pd.read_csv ("manage/repair.csv")
            #    repairString = repairDf.to_json(orient = 'records')
            #    repairJson = json.loads(repairString)
            #    #print(repairJson)
            #    #print(jsonify(repairJson))
            #return jsonify(repairJson)
    except Exception:
        print(traceback.format_exc())

# 刪除指定維修的資料
@app.route('/deleteRepair/<path:filename>', methods = ['GET'])
def deleteRepair(filename):
    try:
        # 更新最新維修狀態
        if os.path.exists(os.path.join("manage/repair.csv")):
            repairDf = pd.read_csv ("manage/repair.csv")
            repairDf = repairDf[~repairDf.nbNumber.str.contains(filename)]
            repairDf.to_csv("manage/repair.csv",encoding='utf_8_sig',index=False)
        response = {"status": 'success'}
        return jsonify(response)
    except Exception:
        print(traceback.format_exc())
        response = {"status": 'failure'}
        return jsonify(response)

# 接收設備借用頁面上傳的資料
@app.route('/rentRecord', methods = ['POST'])
def rent_record():
    try:
        if request.method == 'POST':
            #datename = 'rent'+datetime.today().strftime('%Y-%m-%d')+'.csv'
            body = request.get_data(as_text=True)  # 取得收到的訊息內容
            json_data = json.loads(body) # json 格式化訊息內容
            json_data["parts"] = '、'.join(json_data["parts"])
            json_data["problems"] = '、'.join(json_data["problems"])
            #print(json_data)
            df = pd.json_normalize(json_data)
            # 更新最新借用狀態
            if os.path.exists(os.path.join("manage/rent.csv")):
                rentDf = pd.read_csv ("manage/rent.csv")
                # 只保留最新一筆
                rentDf = rentDf[~rentDf.nbNumber.str.contains(json_data['nbNumber'])] # 移除符合條件的row data
                result = pd.concat([rentDf, df])
                result.to_csv("manage/rent.csv",encoding='utf_8_sig',index=False)
            else:
                df.to_csv("manage/rent.csv",encoding='utf_8_sig',index=False)
            # 儲存借用歷史
            if os.path.exists(os.path.join("manage/history.csv")):
                rentDf = pd.read_csv ("manage/history.csv")
                result = pd.concat([rentDf, df])
                result.to_csv("manage/history.csv",encoding='utf_8_sig',index=False)
            else:
                df.to_csv("manage/history.csv",encoding='utf_8_sig',index=False)
            response = {"status": 'success'}
            return jsonify(response)
    except Exception:
        print(traceback.format_exc())
        
# 列出所有pc清單
@app.route('/info', methods = ['GET'])
def nb_info():
    try:
        if request.method == 'GET':
            # 整理所有管理表單
            manageData = {} # 以筆電編號當 key
            if os.path.exists(os.path.join("manage/rent.csv")):
                with open(os.path.join("manage/rent.csv"),"r",encoding="utf-8") as f:
                    lines = f.readlines()
                    for line in lines:
                        if line.split(",")[4] == "nbNumber":
                            continue
                        manageData[line.split(",")[4]] = line.split(",")
            #print(manageData)

            # 整理所有上傳資料
            nbList = glob.glob("data/*_info.txt")

            # ------------------------------以上傳資料為主去走訪-------------------------------
            nameList = []
            for nb in nbList:
                nameList.append(ntpath.basename(nb.split("_info")[0])) # 取得所有有上傳資料的電腦名稱(=筆電編號 nbNumber)
            response = []
            for name in nameList: # 走訪每台電腦取得所需資訊
                info = {}
                info["nbName"] = name
                info["nbNumber"] = name
                with open(os.path.join("data/", name + '_info.txt')) as f:
                    lines = f.readlines()
                    values = lines[1].split(",")
                    info["os"] =  values[13].replace('"','') + ' ' + values[14].replace('"','') # 取得硬體資訊
                if os.path.exists(os.path.join("data/", name + '_logon.txt')):
                    with open(os.path.join("data/", name + '_logon.txt')) as f:
                        lines = f.readlines()
                        user = ""
                        time = 0
                        for line in lines:
                            if len(line.split()) <= 6:
                                continue
                            else:
                                user = line.split()[0].strip()
                                user = user[1:len(user)]
                                time = line.split()[5].strip()+' '+line.split()[6].strip()+' '+line.split()[7].strip()
                        info["user"] = user # 取得最後登入者
                        info["logonTime"] = time # 取得最後登入時間
                if os.path.exists(os.path.join("data/", name + '_product.txt')):
                    productList = []
                    with open(os.path.join("data/", name + '_product.txt'), encoding='utf-16') as f:
                        lines = f.readlines()
                        for line in lines:
                            if ('Name' in line) and ('Version' in line):
                                continue
                            if len(line.split(",")) == 3:
                                if line.split(",")[1] and line.split(",")[2]:
                                    item = {}
                                    item["name"] = line.split(",")[1]
                                    item["version"] = line.split(",")[2].replace('\n',"").replace('\t',"")
                                    productList.append(item["name"]+":"+item["version"])       
                        info["product"] = productList # 取得軟體清單:版本資訊

                if info["nbName"] in manageData.keys(): # 整合借用資料
                    info["fab"] = manageData[info["nbName"]][0] if manageData[info["nbName"]][0] else '-'
                    info["dep"] = manageData[info["nbName"]][1] if manageData[info["nbName"]][1] else '-'
                    info["employeeId"] = manageData[info["nbName"]][2] if manageData[info["nbName"]][2] else '-'
                    info["name"] = manageData[info["nbName"]][3] if manageData[info["nbName"]][3] else '-'
                    info["parts"] = manageData[info["nbName"]][5] if manageData[info["nbName"]][5] else '-'
                    info["action"] = manageData[info["nbName"]][6] if manageData[info["nbName"]][6] else '-'
                    info["registerTime"] = manageData[info["nbName"]][7] if manageData[info["nbName"]][7] else '-'
                    info["problems"] = manageData[info["nbName"]][8] if manageData[info["nbName"]][8] else '-'
                else:
                    info["fab"] = '-'
                    info["dep"] = '-'
                    info["employeeId"] = '-'
                    info["name"] = '-'
                    info["parts"] = '-'
                    info["action"] = '-'
                    info["registerTime"] = '-'
                    info["problems"] = '-'
                response.append(info)
            
            # ------------------------------以借用資料為主去走訪-------------------------------
            #print(manageData.keys())
            #print(nameList)
            for name in manageData.keys():
                if not name in nameList: # 補上有借用資料但沒上傳資料的筆電
                    #print(name)
                    info = {}
                    info["nbName"] = name
                    info["fab"] = manageData[info["nbName"]][0]
                    info["dep"] = manageData[info["nbName"]][1]
                    info["employeeId"] = manageData[info["nbName"]][2]
                    info["name"] = manageData[info["nbName"]][3]
                    info["nbNumber"] = manageData[info["nbName"]][4]
                    info["parts"] = manageData[info["nbName"]][5]
                    info["action"] = manageData[info["nbName"]][6]
                    info["registerTime"] = manageData[info["nbName"]][7]
                    info["problems"] = manageData[info["nbName"]][8]
                    info["user"] = '-' # 最後登入者
                    info["logonTime"] = '-' # 最後登入時間
                    info["os"] = '-' # 廠牌型號
                    response.append(info)
            return jsonify(response)
    except Exception:
        print(traceback.format_exc())

# 找尋有安裝某特定軟體的pc清單
@app.route('/search/<path:product>', methods = ['GET'])
def searchProduct(product):
    try:
        if request.method == 'GET':
            # 整理所有上傳資料
            nbList = glob.glob("data/*_info.txt")
            nameList = []
            for nb in nbList:
                nameList.append(ntpath.basename(nb.split("_info")[0])) # 取得所有電腦名稱
            response = []
            for name in nameList:
                info = {}
                info["name"] = name
                with open(os.path.join("data/", name + '_info.txt')) as f:
                    lines = f.readlines()
                    values = lines[1].split(",")
                    info["os"] = values[1].replace('"','')
                if os.path.exists(os.path.join("data/", name + '_product.txt')):
                    productList = []
                    with open(os.path.join("data/", name + '_product.txt'), encoding='utf-16') as f:
                        lines = f.readlines()
                        for line in lines:
                            if ('Name' in line) and ('Version' in line):
                                continue
                            if len(line.split(",")) == 3:
                                if line.split(",")[1] and line.split(",")[2]:
                                    item = {}
                                    if product in line.split(",")[1]:
                                        item["name"] = line.split(",")[1]
                                        item["version"] = line.split(",")[2].replace('\n',"").replace('\t',"")
                                        productList.append(item)
                        info["product"] = productList
                if os.path.exists(os.path.join("data/", name + '_logon.txt')):
                    with open(os.path.join("data/", name + '_logon.txt')) as f:
                        lines = f.readlines()
                        user = ""
                        time = 0
                        for line in lines:
                            if len(line.split()) <= 6:
                                continue
                            else:
                                user = line.split()[0].strip()
                                user = user[1:len(user)]
                                time = line.split()[5].strip()+' '+line.split()[6].strip()+' '+line.split()[7].strip()
                        info["user"] = user
                        info["time"] = time
                response.append(info)
            return jsonify(response)
    except Exception:
        print(traceback.format_exc())

# 下載單pc詳細資訊
@app.route('/download/<path:filename>', methods = ['GET'])
def downloadFile(filename):
    try:
        wb = openpyxl.Workbook()
        # 建立設備資訊頁籤
        sheet1 = wb.create_sheet("設備資訊", 0)
        if os.path.exists(os.path.join("data/", filename + '_info.txt')):
            with open(os.path.join("data/", filename + '_info.txt')) as f:
                lines = f.readlines()
                headers = lines[0].split(",")
                values = lines[1].split(",")
                for idx, val in enumerate(headers):
                    sheet1.append((headers[idx].replace('"',''), values[idx].replace('"','')))
        # 建立軟體清單頁籤
        sheet2 = wb.create_sheet("軟體清單", 1)
        sheet2.append(("軟體名稱","版本"))
        if os.path.exists(os.path.join("data/", filename + '_product.txt')):
            with open(os.path.join("data/", filename + '_product.txt'), encoding='utf-16') as f:
                lines = f.readlines()
                for line in lines:
                    if ('Name' in line) and ('Version' in line):
                        continue
                    if len(line.split(",")) == 3:
                        if line.split(",")[1] and line.split(",")[2]:
                            sheet2.append((line.split(",")[1], line.split(",")[2].replace('\n',"").replace('\t',"")))
        # 建立者登入時間頁籤
        sheet3 = wb.create_sheet("使用者登入時間", 2)
        sheet3.append(("使用者","登入時間"))
        if os.path.exists(os.path.join("data/", filename + '_logon.txt')):
            with open(os.path.join("data/", filename + '_logon.txt')) as f:
                lines = f.readlines()
                for line in lines:
                    if len(line.split()) <= 6:
                        continue
                    else:
                        user = line.split()[0].strip()
                        user = user[1:len(user)]
                        time = line.split()[5].strip()+' '+line.split()[6].strip()+' '+line.split()[7].strip()
                    sheet3.append((user, time))
        wb.save(os.path.join("data/", filename+'.xlsx'))
        return send_file(os.path.join("data/", filename+'.xlsx'), as_attachment=True)
    except Exception:
        print(traceback.format_exc())

# 刪除指定pc的資料
@app.route('/deleteInfo/<path:filename>', methods = ['GET'])
def deleteFile(filename):
    try:
        if os.path.exists(os.path.join("data/", filename + '_info.txt')):
            os.remove(os.path.join("data/", filename + '_info.txt'))
        if os.path.exists(os.path.join("data/", filename + '_product.txt')):
            os.remove(os.path.join("data/", filename + '_product.txt'))
        if os.path.exists(os.path.join("data/", filename + '_logon.txt')):
            os.remove(os.path.join("data/", filename + '_logon.txt'))
        if os.path.exists(os.path.join("data/", filename + '.xlsx')):
            os.remove(os.path.join("data/", filename + '.xlsx'))
        # 更新最新借用狀態
        if os.path.exists(os.path.join("manage/rent.csv")):
            rentDf = pd.read_csv ("manage/rent.csv")
            rentDf = rentDf[~rentDf.nbNumber.str.contains(filename)]
            rentDf.to_csv("manage/rent.csv",encoding='utf_8_sig',index=False)
        response = {"status": 'success'}
        return jsonify(response)
    except Exception:
        print(traceback.format_exc())
        response = {"status": 'failure'}
        return jsonify(response)

# 接收要匯出的設備管理清單資訊
@app.route('/export', methods = ['POST'])
def export():
    try:
        if request.method == 'POST':
            print(request.get_json())
            wb = openpyxl.Workbook()
            # 建立設備資訊頁籤
            sheet1 = wb.create_sheet("設備借用狀態", 0)
            sheet1.append(('廠', '部門', '工號', '姓名', '筆電編號', '最後登入帳號', 
            '作業系統', '最近開機時間', '借用日期', '狀態', '應用程式名稱', '應用程式版本'))
            for row in request.get_json():
                productName = ''
                productVersion = ''
                if (len(row['production'].split(':')) == 2): 
                    productName = row['production'].split(':')[0]
                    productVersion = row['production'].split(':')[1]
                
                sheet1.append((row['fab'], row['dep'], row['employeeId'], row['name'], 
                        row['nbNumber'], row['user'], row['os'], row['logonTime'], 
                        row['borrowTime'], row['action'], productName, productVersion))
        wb.save(os.path.join("data/", '設備借用狀態.xlsx'))
        response = {"status": 'success'}
        return jsonify(response)
    except Exception:
        print(traceback.format_exc())

# 匯出的設備管理清單資訊
@app.route('/exportSummary', methods = ['GET'])
def exportSummary():
    return send_file(os.path.join("data/", '設備借用狀態.xlsx'), as_attachment=True)

# chart1Option api
@app.route('/chart1Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart1Option(startDate, endDate):
    d0 = datetime.strptime(startDate, '%Y-%m-%d')
    d1 = datetime.strptime(endDate, '%Y-%m-%d')
    delta = d1 - d0
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/history.csv")):
                df = pd.read_csv ("manage/history.csv")
                df["registerTime"] = pd.to_datetime(df["registerTime"], format="%Y/%m/%d %H:%M:%S")  #轉換日期時間格式(要計算時間)
                if delta.days <= 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m-%d') #從datetime取出日期到date欄位
                elif delta.days > 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m')

                group_count = df.groupby(["date", "action"], as_index=False)["name"].count() #先分組
                group_count = group_count.rename(columns={"name": "count"}, inplace=False)
                group_count = group_count.sort_values(["date"])
                group_count.index = range(len(group_count)) #重新排序index
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

# chart2Option api
@app.route('/chart2Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart2Option(startDate, endDate):
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/history.csv")):
                df = pd.read_csv ("manage/history.csv")
                group_count = df.groupby(["action"], as_index=False)["name"].count() #先分組
                group_count = group_count.rename(columns={"name": "count"}, inplace=False)
                group_count.index = range(len(group_count)) #重新排序index
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

# chart3Option api
@app.route('/chart3Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart3Option(startDate, endDate):
    d0 = datetime.strptime(startDate, '%Y-%m-%d')
    d1 = datetime.strptime(endDate, '%Y-%m-%d')
    delta = d1 - d0
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/history.csv")):
                df = pd.read_csv ("manage/history.csv")
                df["registerTime"] = pd.to_datetime(df["registerTime"], format="%Y/%m/%d %H:%M:%S")  #轉換日期時間格式(要計算時間)
                if delta.days <= 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m-%d') #從datetime取出日期到date欄位
                elif delta.days > 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m')
                
                group_count1 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('藍屏')].count()).reset_index(name="count")
                group_count1['problem'] = '藍屏'
                group_count2 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('黑屏')].count()).reset_index(name="count")
                group_count2['problem'] = '黑屏'
                group_count3 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('重開')].count()).reset_index(name="count")
                group_count3['problem'] = '重開'
                group_count4 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('WIFI連不上')].count()).reset_index(name="count")
                group_count4['problem'] = 'WIFI連不上'
                group_count5 = df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('其它')].count()).reset_index(name="count")
                group_count5['problem'] = '其它'
                group_count = pd.concat([group_count1, group_count2, group_count3, group_count4, group_count5])
                group_count = group_count.sort_values(["date"])
                group_count.index = range(len(group_count)) #重新排序index
                #print(group_count)
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

#chart4Option api
@app.route('/chart4Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart4Option(startDate, endDate):
    d0 = datetime.strptime(startDate, '%Y-%m-%d')
    d1 = datetime.strptime(endDate, '%Y-%m-%d')
    delta = d1 - d0
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/history.csv")):
                df = pd.read_csv ("manage/history.csv")
                df["registerTime"] = pd.to_datetime(df["registerTime"], format="%Y/%m/%d %H:%M:%S")  #轉換日期時間格式(要計算時間)
                if delta.days <= 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m-%d') #從datetime取出日期到date欄位
                elif delta.days > 7:
                    df["date"] = df["registerTime"].dt.strftime('%Y-%m')

                group_count1 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('藍屏')].count()).reset_index(name="count")
                group_count1['problem'] = '藍屏'
                group_count2 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('黑屏')].count()).reset_index(name="count")
                group_count2['problem'] = '黑屏'
                group_count3 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('重開')].count()).reset_index(name="count")
                group_count3['problem'] = '重開'
                group_count4 =  df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('WIFI連不上')].count()).reset_index(name="count")
                group_count4['problem'] = 'WIFI連不上'
                group_count5 = df.groupby(['date'])['problems'].apply(lambda x: x[x.str.contains('其它')].count()).reset_index(name="count")
                group_count5['problem'] = '其它'
                group_count = pd.concat([group_count1, group_count2, group_count3, group_count4, group_count5])
                group_count = group_count.groupby(["problem"], as_index=False)["count"].sum()
                group_count.index = range(len(group_count)) #重新排序index
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

# chart5Option api
@app.route('/chart5Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart5Option(startDate, endDate):
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/rent.csv")):
                df = pd.read_csv ("manage/rent.csv")
                group_count = df.groupby(["fab"], as_index=False)["action"].apply(lambda x: x[x.str.contains('歸還')].count())
                group_count = group_count.rename(columns={"action": "count"}, inplace=False)
                group_count.index = range(len(group_count)) #重新排序index
                #print(group_count)
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

# chart6Option api
@app.route('/chart6Option/<path:startDate>/<path:endDate>', methods = ['GET'])
def chart6Option(startDate, endDate):
    try:
        if request.method == 'GET':
            response = {'data': []}
            if os.path.exists(os.path.join("manage/repair.csv")):
                df = pd.read_csv ("manage/repair.csv")
                group_count = df.groupby(["fab"], as_index=False)["repairStatus"].apply(lambda x: x[x.str.contains('repairing')].count())
                group_count = group_count.rename(columns={"repairStatus": "count"}, inplace=False)
                group_count.index = range(len(group_count)) #重新排序index
                #print(group_count)
                response["data"] = json.loads(group_count.to_json(orient = 'records'))
                #print(response)
            return response
    except Exception:
        print(traceback.format_exc())

@app.route("/auth", methods = ['POST'])
def auth():
    try:
        if request.method == 'POST':
            #print(request.form)
            username = request.form["username"]
            password = request.form["password"]
            if 'admin' in username and 'admin' in password:
                response = {"status": 'success'}
                return jsonify(response)
            else:
                response = {"status": 'failure'}
                return jsonify(response)
    except Exception:
        print(traceback.format_exc())

@app.route("/login")
def login():
    return app.send_static_file('pages/manage/login.html')

@app.route("/")
def index():
    return app.send_static_file('pages/manage/borrow.html')

@app.route("/device")
def device():
    return app.send_static_file('pages/manage/device.html')

@app.route("/repair")
def repair():
    return app.send_static_file('pages/manage/repair.html')

@app.route("/analysis")
def analysis():
    return app.send_static_file('pages/manage/analysis.html')

if __name__ == '__main__':
#    app.run(host='0.0.0.0', port=config["DEFAULT"]["API_PORT"])
   app.run(host='0.0.0.0', port=80)